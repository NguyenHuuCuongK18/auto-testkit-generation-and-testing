import wx
import subprocess
import threading
import os
import openpyxl
import time
import json
from datetime import datetime
from threading import Lock
from typing import List, Optional
from gooey import Gooey, GooeyParser
import wx.adv  # For wx.MessageDialog

# Global lock for thread-safe file writes
_file_write_lock: Lock = Lock()

class InteractiveFrame(wx.Frame):
    """GUI frame for generating test cases for client-server applications."""
    def __init__(self, args: GooeyParser) -> None:
        """
        Initialize the InteractiveFrame with provided arguments.

        Args:
            args (GooeyParser): Parsed command-line arguments containing client/server paths,
                               test case name, and save location.
        """
        super().__init__(None, title=f"Test Case Generator - {args.test_case_name}", size=(800, 600))
        self.args: GooeyParser = args
        self.current_stage: int = 1
        self.inputs: List[str] = []
        self.client_process: Optional[subprocess.Popen] = None
        self.server_process: Optional[subprocess.Popen] = None
        self.thread_client: Optional[threading.Thread] = None
        self.thread_server: Optional[threading.Thread] = None
        self.client_record_file: Optional[str] = None
        self.server_record_file: Optional[str] = None
        self.client_output: List[str] = []  # Store client console output
        self.server_output: List[str] = []  # Store server console output
        self.excel_file: Optional[str] = None
        self.meta_file: Optional[str] = None

        # Initialize UI elements
        self.console: wx.TextCtrl = wx.TextCtrl(self, style=wx.TE_MULTILINE | wx.TE_READONLY | wx.HSCROLL)
        self.stage_label: wx.StaticText = wx.StaticText(self, label=f"Current Stage: {self.current_stage}")
        self.input_label: wx.StaticText = wx.StaticText(self, label="Enter input (number or value):")
        self.input_entry: wx.TextCtrl = wx.TextCtrl(self, style=wx.TE_PROCESS_ENTER)
        self.input_entry.Bind(wx.EVT_TEXT_ENTER, self.on_submit)
        submit_btn: wx.Button = wx.Button(self, label="Submit Input")
        submit_btn.Bind(wx.EVT_BUTTON, self.on_submit)
        record_btn: wx.Button = wx.Button(self, label="Record")
        record_btn.Bind(wx.EVT_BUTTON, self.on_record)
        end_btn: wx.Button = wx.Button(self, label="End All Processes")
        end_btn.Bind(wx.EVT_BUTTON, self.on_end_processes)

        # Set up layout
        sizer: wx.BoxSizer = wx.BoxSizer(wx.VERTICAL)
        sizer.Add(self.console, 1, wx.EXPAND | wx.ALL, 5)
        sizer.Add(self.stage_label, 0, wx.ALL, 5)
        sizer.Add(self.input_label, 0, wx.ALL, 5)
        sizer.Add(self.input_entry, 0, wx.EXPAND | wx.ALL, 5)
        btn_sizer: wx.BoxSizer = wx.BoxSizer(wx.HORIZONTAL)
        btn_sizer.Add(submit_btn, 0, wx.ALL, 5)
        btn_sizer.Add(record_btn, 0, wx.ALL, 5)
        btn_sizer.Add(end_btn, 0, wx.ALL, 5)
        sizer.Add(btn_sizer, 0, wx.ALIGN_CENTER | wx.ALL, 5)
        self.SetSizer(sizer)

        # Bind close event
        self.Bind(wx.EVT_CLOSE, self.on_close)

        # Initialize folders and processes
        self.setup_folders()
        self.start_processes()
        self.Show()

    def setup_folders(self) -> None:
        """Set up the folder structure for saving test case data and initialize record files."""
        test_case_folder: str = os.path.join(self.args.save_location, self.args.test_case_name)
        if os.path.exists(test_case_folder):
            dlg: wx.MessageDialog = wx.MessageDialog(
                self, f"Test case folder '{self.args.test_case_name}' already exists. Overwrite?", "Confirm", wx.YES_NO
            )
            if dlg.ShowModal() != wx.ID_YES:
                self.Close()
                return
        os.makedirs(test_case_folder, exist_ok=True)
        record_folder: str = os.path.join(test_case_folder, "record")
        os.makedirs(record_folder, exist_ok=True)
        self.client_record_file = os.path.join(record_folder, "client_record.txt")
        self.server_record_file = os.path.join(record_folder, "server_record.txt")
        self.excel_file = os.path.join(test_case_folder, "testcase.xlsx")
        self.meta_file = os.path.join(test_case_folder, "meta.json")

        with _file_write_lock:
            open(self.client_record_file, 'w', encoding='utf-8', errors='replace').close()
            open(self.server_record_file, 'w', encoding='utf-8', errors='replace').close()

    def start_processes(self) -> None:
        """Start client and server processes and their output reading threads."""
        client_path: str = self.args.client_path
        server_path: str = self.args.server_path

        if not os.path.exists(client_path) or not os.path.exists(server_path):
            wx.MessageBox("Selected executable files do not exist.", "Error")
            self.Close()
            return

        try:
            self.server_process = subprocess.Popen(
                server_path,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                stdin=subprocess.PIPE,
                text=True,
                bufsize=1,
                universal_newlines=True
            )
            time.sleep(1.2)  # Allow server to initialize
            self.client_process = subprocess.Popen(
                client_path,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                stdin=subprocess.PIPE,
                text=True,
                bufsize=1,
                universal_newlines=True
            )
            self.append_to_console("Processes started successfully.\n")
        except Exception as e:
            wx.MessageBox(f"Failed to start processes: {e}", "Error")
            self.Close()
            return

        self.client_output = []  # Reset outputs
        self.server_output = []
        self.thread_server = threading.Thread(target=self.read_output, args=(self.server_process, self.server_output))
        self.thread_client = threading.Thread(target=self.read_output, args=(self.client_process, self.client_output))
        self.thread_server.daemon = True
        self.thread_client.daemon = True
        self.thread_server.start()
        self.thread_client.start()

    def read_output(self, process: subprocess.Popen, output_list: List[str]) -> None:
        """
        Read output from a process and append it to the output list.

        Args:
            process (subprocess.Popen): The process to read output from.
            output_list (List[str]): The list to store the output lines.
        """
        while True:
            try:
                line: str = process.stdout.readline()
                if not line:
                    break
                normalized_line: str = line.replace('\r\n', '\n').rstrip('\n') + '\n'
                with _file_write_lock:
                    output_list.append(normalized_line)
                wx.CallAfter(self.append_to_console, normalized_line)
            except Exception:
                break

    def append_to_console(self, text: str) -> None:
        """
        Append text to the console display.

        Args:
            text (str): The text to append.
        """
        self.console.AppendText(text)

    def on_end_processes(self, event: wx.Event) -> None:
        """Handle the 'End All Processes' button click."""
        self.append_to_console("Ending all processes...\n")
        self.cleanup_processes()
        self.append_to_console("Processes ended. You can restart by submitting a new input.\n")

    def on_submit(self, event: wx.Event) -> None:
        """Handle input submission, send to client process, and update stage."""
        # Restart processes if they are not running
        if self.client_process is None or self.client_process.poll() is not None or \
           self.server_process is None or self.server_process.poll() is not None:
            self.append_to_console("Processes not running. Restarting...\n")
            self.start_processes()
            time.sleep(2.0)  # Allow processes to restart

        value: str = self.input_entry.GetValue().strip()
        if not value:
            wx.MessageBox("Input is required.", "Error")
            return
        self.inputs.append(value)
        self.append_to_console(f"[Input] Stage {self.current_stage}: {value}\n")
        try:
            self.client_process.stdin.write(value + '\n')
            self.client_process.stdin.flush()
        except Exception as e:
            wx.MessageBox(f"Failed to send input to client: {e}", "Error")
            self.cleanup_processes()
            return
        self.input_entry.Clear()
        self.current_stage += 1
        self.stage_label.SetLabel(f"Current Stage: {self.current_stage}")
        time.sleep(0.5)  # Allow output capture

    def on_record(self, event: wx.Event) -> None:
        """Record test case data to files and create Excel and meta files."""
        if self.client_process is None or self.client_process.poll() is not None or \
           self.server_process is None or self.server_process.poll() is not None:
            wx.MessageBox("Processes must be running to record. Please submit an input to start them.", "Error")
            return

        time.sleep(0.6)  # Allow output to settle
        dlg: wx.TextEntryDialog = wx.TextEntryDialog(self, "Enter points for this test case:", "Enter Points")
        if dlg.ShowModal() == wx.ID_OK:
            points: str = dlg.GetValue().strip()
            if not points:
                wx.MessageBox("Points are required.", "Error")
                return

            time.sleep(3.0)  # Allow final outputs to be captured

            self.append_to_console("Recording test case...\n")

            # Write outputs to record files
            with _file_write_lock:
                with open(self.client_record_file, 'w', encoding='utf-8', errors='replace') as f:
                    f.writelines(self.client_output)
                with open(self.server_record_file, 'w', encoding='utf-8', errors='replace') as f:
                    f.writelines(self.server_output)

            # Create meta file
            meta: dict = {
                "test_case_name": self.args.test_case_name,
                "stages": len(self.inputs),
                "inputs": self.inputs,
                "points": points,
                "timestamp": datetime.now().isoformat()
            }
            with open(self.meta_file, 'w', encoding='utf-8') as f:
                json.dump(meta, f, indent=4, ensure_ascii=False)

            # Create Excel file
            wb: openpyxl.Workbook = openpyxl.Workbook()
            for stage in range(1, self.current_stage):
                ws: openpyxl.worksheet.worksheet.Worksheet = wb.create_sheet(title=f"Stage {stage}")
                for i in range(1, stage + 1):
                    ws.cell(row=i, column=1).value = f"Stage {i}"
                    ws.cell(row=i, column=2).value = "input"
                    ws.cell(row=i, column=3).value = self.inputs[i - 1] if i - 1 < len(self.inputs) else ""
                if stage == self.current_stage - 1:
                    ws.cell(row=1, column=4).value = "record"
                    for i in range(1, stage + 1):
                        ws.cell(row=i, column=4).value = "yes" if i == stage else "none"
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            wb.save(self.excel_file)

            wx.MessageBox(f"Recorded to {self.excel_file}, {os.path.dirname(self.client_record_file)}, and {self.meta_file}", "Success")
            self.Close()

    def cleanup_processes(self) -> None:
        """Clean up running processes and threads, resetting inputs and stage."""
        if self.client_process and self.client_process.poll() is None:
            self.client_process.terminate()
            try:
                self.client_process.wait(timeout=3)
            except subprocess.TimeoutExpired:
                self.client_process.kill()
        if self.server_process and self.server_process.poll() is None:
            self.server_process.terminate()
            try:
                self.server_process.wait(timeout=3)
            except subprocess.TimeoutExpired:
                self.server_process.kill()
        if self.thread_client:
            self.thread_client.join(timeout=5)
        if self.thread_server:
            self.thread_server.join(timeout=5)
        self.client_process = None
        self.server_process = None
        self.thread_client = None
        self.thread_server = None
        self.inputs = []
        self.current_stage = 1
        self.stage_label.SetLabel(f"Current Stage: {self.current_stage}")

    def on_close(self, event: wx.Event) -> None:
        """
        Handle window close event, ensuring processes are cleaned up.

        Args:
            event (wx.Event): The close event.
        """
        self.append_to_console("Closing application. Cleaning up processes...\n")
        self.cleanup_processes()
        event.Skip()

@Gooey(program_name="Test Case Generator", default_size=(800, 600))
def main() -> None:
    """Main function to set up the GUI and parse command-line arguments."""
    parser: GooeyParser = GooeyParser(description="Generate test cases for client-server applications")
    parser.add_argument('client_path', help="Client Executable", widget='FileChooser')
    parser.add_argument('server_path', help="Server Executable", widget='FileChooser')
    parser.add_argument('test_case_name', help="Test Case Name")
    parser.add_argument('save_location', help="Save Location (directory)", widget='DirChooser', default=os.getcwd())
    args: GooeyParser = parser.parse_args()

    app: wx.App = wx.App(False)
    frame: InteractiveFrame = InteractiveFrame(args)
    app.MainLoop()

if __name__ == '__main__':
    main()